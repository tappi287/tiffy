from openpyxl import load_workbook, worksheet
from openpyxl.utils import column_index_from_string
from modules.exiftool import Exif
from modules.log import init_logging

LOGGER = init_logging(__name__)


class ReadExcel:
    @classmethod
    def get_data(cls, file):
        ws = cls.load(file)
        return cls.read_worksheet(ws)

    @classmethod
    def load(cls, file) -> worksheet:
        wb = load_workbook(file)
        return wb.active

    @classmethod
    def read_worksheet(cls, ws):
        # Find start row behind header
        for start_row in range(ws.min_row, ws.max_row):
            row_value = ws.cell(start_row, ws.min_column).value
            if row_value:
                break
        else:
            start_row = 1

        excel_data = dict()

        for row in range(start_row, ws.max_row):
            file_name = ws[f'{Exif.spreadsheet_file_column}{row}'].value
            excel_data[file_name] = dict()

            for result in cls._read_columns(row, ws):
                key, value = result
                excel_data[file_name].update({key: value})

            LOGGER.debug('%s - %s', file_name, excel_data[file_name])

        return excel_data

    @staticmethod
    def _read_columns(row, ws):
        for sheet_map in Exif.spreadsheet_map.items():
            key, column = sheet_map
            if not column:
                continue
            col = column_index_from_string(column)

            yield key, ws.cell(row, col).value
