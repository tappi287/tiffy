from pathlib import Path
from time import time
from openpyxl import load_workbook, worksheet
from openpyxl.utils import column_index_from_string
from PyQt5.QtCore import QThread, QObject, pyqtSignal
from PyQt5.QtWidgets import QTreeWidgetItem

from modules import TiffySettings
from modules.exif_worker import Exif
from modules.widgets.file_dialog import FileDialog
from modules.detect_language import get_translation
from modules.log import init_logging

LOGGER = init_logging(__name__)

# translate strings
lang = get_translation()
lang.install()
_ = lang.gettext


class OpenExcel(QObject):
    def __init__(self, ui, menu):
        super(OpenExcel, self).__init__(ui)
        self.ui, self.menu = ui, menu
        self.progress = ui.progress_widget.progress
        self.thread = None
        self.load_timer = time()

    def open_excel(self, file: str=None) -> None:
        self.menu.enable_menus(False)

        if not file:
            file = FileDialog.open(self.ui, None, 'xlsx')

        if not file:
            LOGGER.info('Open Xlsx File dialog canceled.')
            self.menu.enable_menus(True)
            return

        TiffySettings.add_recent_file(file, 'xlsx')

        self.thread = ExcelThread(self, file)
        self.thread.result.connect(self.excel_result)
        self.thread.start()
        self.load_timer = time()

    def excel_result(self, excel_data):
        self.progress.hide()

        if not excel_data:
            LOGGER.info('Xlsx returned no data.')
            self.menu.enable_menus(True)
            return

        self.ui.treeWidget.clear()
        file = ''

        for data in excel_data.items():
            file, file_dict = data
            item_list = [f'{file}']

            for tag, value in zip(Exif.exiftool_tags, file_dict.values()):
                if not value:
                    value = _('[Keine Daten]')
                item_list.append(value)

            item = QTreeWidgetItem(item_list)
            self.ui.treeWidget.addTopLevelItem(item)

        # Store excel data in Ui class
        self.ui.excel_data = excel_data

        # Output load info
        load_time = time() - self.load_timer
        self.ui.statusBar().showMessage(_('Excel Tabelle in {0:.3}s geladen.').format(load_time), 10000)
        self.menu.enable_menus(True)

    def setup_progress(self, value):
        self.progress.setValue(0)
        self.progress.setMaximum(value)
        self.progress.show()

    def update_progress(self):
        self.progress.setFormat('%v / %m')
        v = self.progress.value() + 1
        self.progress.setValue(v)


class ExcelThread(QThread):
    result = pyqtSignal(dict)

    def __init__(self, parent, file):
        super(ExcelThread, self).__init__(parent)
        self.parent, self.file = parent, file

    def run(self):
        excel_reader = ReadExcel(self.parent)
        excel_reader.progress.connect(self.parent.update_progress)
        excel_reader.num_items.connect(self.parent.setup_progress)
        excel_reader.num_items.emit(0)
        excel_data = excel_reader.get_data(Path(self.file).as_posix())

        self.result.emit(excel_data)
        LOGGER.debug('Excel reader thread ended.')


class ReadExcel(QObject):
    progress = pyqtSignal()
    num_items = pyqtSignal(int)

    def __init__(self, parent):
        super(ReadExcel, self).__init__(parent)

    def get_data(self, file):
        ws = self.load(file)
        return self.read_worksheet(ws)

    def read_worksheet(self, ws):
        # Find start row behind header
        for start_row in range(ws.min_row, ws.max_row):
            row_value = ws.cell(start_row, ws.min_column).value
            if row_value:
                break
        else:
            start_row = 1

        self.num_items.emit(ws.max_row - start_row)
        excel_data = dict()
        file_name = ''

        for row in range(start_row, ws.max_row + 1):
            for result in self._read_columns(row, ws):
                key, value = result

                if key == 'file':
                    file_name = value
                    excel_data[file_name] = dict()
                    continue

                if not file_name:
                    continue

                excel_data[file_name].update({key: value})

            LOGGER.debug('%s - %s', file_name, excel_data[file_name])
            self.progress.emit()

        return excel_data

    @staticmethod
    def load(file) -> worksheet:
        wb = load_workbook(file)
        return wb.active

    @staticmethod
    def _read_columns(row, ws):
        for sheet_map in Exif.spreadsheet_map.items():
            key, column = sheet_map

            if not column:
                yield key, None
                continue

            col = column_index_from_string(column)

            yield key, ws.cell(row, col).value
