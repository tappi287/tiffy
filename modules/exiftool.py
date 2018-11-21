import exiftool

from pathlib import Path
from openpyxl import load_workbook, worksheet
from openpyxl.utils import column_index_from_string
from PyQt5 import QtCore, QtWidgets

from modules.app_globals import EXIFTOOL_BINARY
from modules.detect_language import get_translation
from modules.log import init_logging

LOGGER = init_logging(__name__)

# translate strings
lang = get_translation()
lang.install()
_ = lang.gettext


class Exif(QtCore.QObject):
    result = QtCore.pyqtSignal(str, str)
    file_types = ['.tif', '.tiff']

    xmp_keys = ['XMP:Title', 'XMP:Creator', 'XMP:Description', 'XMP:Subject', 'XMP:Rights']
    iptc_keys = ['IPTC:ObjectName', 'IPTC:By-line', 'IPTC:Caption-Abstract', 'IPTC:Keywords', 'IPTC:CopyrightNotice']
    exif_keys = [None, 'EXIF:Artist', 'EXIF:ImageDescription', None, 'EXIF:Copyright']
    exiftool_tags = ['-Title', '-Creator', '-Description', '-Subject', '-Rights']

    spreadsheet_file_column = 'B'
    spreadsheet_map = {
        'title': 'K',
        'author': None,
        'description': None,
        'keywords': 'B',
        'copyright': 'EG'
        }

    excel_data = None

    def __init__(self, ui: QtWidgets.QMainWindow, img_path: Path):
        super(Exif, self).__init__(ui)
        self.ui, self.img_path = ui, img_path

    def update_meta_data(self, cmd_list, file_name):
        with exiftool.ExifTool(executable_=EXIFTOOL_BINARY) as et:
            LOGGER.debug('Sending commands %s', cmd_list)

            params = map(exiftool.fsencode, cmd_list)
            result = et.execute(*params).decode('UTF-8')

            LOGGER.debug('Exiftool result: %s', result)
            self.result.emit(file_name, result)

    def read_meta_data(self):
        with exiftool.ExifTool(executable_=EXIFTOOL_BINARY) as et:
            metadata = et.get_metadata_batch(self.get_img_files().as_posix())

        if not metadata:
            return

        self.result.emit(metadata)

    def get_img_files(self):
        for img_file in self.img_path.glob('*.*'):
            if f'{img_file.suffix}'.casefold() in self.file_types:
                yield img_file


class ReadExcel:
    @classmethod
    def get_data(cls, file):
        ws = cls.load(file)
        return cls.read_worksheet(ws)

    @classmethod
    def load(cls, file) -> worksheet:
        wb = load_workbook(file)
        return wb.active

    @classmethod
    def read_worksheet(cls, ws):
        # Find start row behind header
        for start_row in range(ws.min_row, ws.max_row):
            row_value = ws.cell(start_row, ws.min_column).value
            if row_value:
                break
        else:
            start_row = 1

        excel_data = dict()

        for row in range(start_row, ws.max_row + 1):
            file_name = ws[f'{Exif.spreadsheet_file_column}{row}'].value
            excel_data[file_name] = dict()

            for result in cls._read_columns(row, ws):
                key, value = result
                excel_data[file_name].update({key: value})

            LOGGER.debug('%s - %s', file_name, excel_data[file_name])

        return excel_data

    @staticmethod
    def _read_columns(row, ws):
        for sheet_map in Exif.spreadsheet_map.items():
            key, column = sheet_map

            if not column:
                yield key, None
                continue

            col = column_index_from_string(column)

            yield key, ws.cell(row, col).value